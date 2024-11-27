from openpyxl import Workbook , load_workbook
wb = Workbook()    #Write data to excel

ws = wb.active
print(wb.sheetnames)
#ws['A1'] = "ABC"

list = [["abc","ggg",123],["nnn","jij",234],["mmm","rrr",542]]

for data in list:
    ws.append(data)

wb.save("Efile2.xlsx")
print(ws.cell(row = 2, column = 2).value)

wr = load_workbook(filename='Efile2.xlsx')  #reading data from excel
sh = wr['Sheet']

print(sh['A1'].value)
row_cont = sh.max_row
col_cont = sh.max_column
print(row_cont)
print(col_cont)

for i in range(1,row_cont+1):
    for j in range(1,col_cont+1):
        print(sh.cell(row=i,column=j).value)